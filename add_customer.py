# Adds new customers to the customer database
from faker import Faker
import openpyxl


def add_customer(n):
    """Creates dummy customers.

    Returns:
        list. List of dictionaries containing customer info.
    """

    f = Faker()
    num_cust = n
    customers = []
    for customer in range(1, num_cust + 1):
        cust_data = {'first_name': f.first_name(), 'last_name': f.last_name(),
                     'phone': f.phone_number(), 'email': f.email(),
                     'date': f.date_time_this_year()}

        customers.append(cust_data)
    return customers


def update_worksheet(customer_list):
    """Adds customer(s) to the database.

    Attributes:
        customer_list: list. A list of dictionaries containing customer info.
    """

    # Write customer details to a database
    wb = openpyxl.load_workbook('customer_data-edited-dates.xlsx')
    sheet = wb.active

    for row in range(len(customer_list)):
        # add 2 to row to account for header row and zero based indexing.
        sheet.cell(row=row + 2, column=1).value = \
            customer_list[row]['first_name']
        sheet.cell(row=row + 2, column=2).value = \
            customer_list[row]['last_name']
        sheet.cell(row=row + 2, column=3).value = \
            customer_list[row]['phone']
        sheet.cell(row=row + 2, column=4).value = \
            customer_list[row]['email']
        sheet.cell(row=row + 2, column=5).value = \
            customer_list[row]['date']
    wb.save('customer_data-edited-dates.xlsx')

if __name__ == "__main__":
    update_worksheet(add_customer(100))
