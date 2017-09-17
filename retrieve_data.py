# Retrieves data from the customer database and writes it to a new spreadsheet
import sys
import argparse
import openpyxl
import config


def parse_args(args):
    parser = argparse.ArgumentParser(
        prog="retrieve_data", description="Retrieve data from a spreadsheet",
        usage='%(prog)s [options]')
    parser.add_argument(
        "days", type=int, help="Retrieve customer data from the last n days",
        choices=[7, 15, 30])
    args = parser.parse_args(args)
    # main(args.days)
    return args.days


def main(n):
    create(parse_args(n))


def get_spreadsheet_data():
    """Reads data from the original customer spreadsheet"""
    try:
        wb = openpyxl.load_workbook(config.customer_data)
        sheet = wb.active
        # To get rows from worksheet
        rows = \
            [row for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row)]
        return rows
    except IOError:
        raise IOError('Customer database not found. Ensure that the customer '
                      'info spreadsheet exists and that the config file points'
                      ' to it')


def string_to_index(col):
    """Returns Cell column index from string"""
    return openpyxl.utils.column_index_from_string(col)


def get_cust_data_from_worksheet(rows, days, date_delta):
    """Retrieves worksheet data in a useful format.

    Attributes:
        rows: list. A list of all the rows retrieved from the spreadsheet.
        date_delta: datetime.datetime. The date range to search for.
    Returns:
        values: list. A list of the records in the last N days where N is the
            date_delta
    """
# Search for records added in the last n days
    values = []
    for row in rows:
        if row[4].value >= date_delta:
            for cell in row:
                values.append(cell.value)
    if len(values) == 0:
        print "No records added in the last {} days".format(days)
        sys.exit()

    # values is a long unordered list, it must be converted into a neat list of
    # tuples that looks like this:
    # [('Name'),('LastName'),('phone'),('email'),('date')]

    new_values_list = zip(*[iter(values)]*5)
    return new_values_list


def create(days):
    """Creates new spreadsheet with customer info added in the last n days

    Attributes:
        days: int. The search query will search for records added in this
             number of days.
    """

    if days == 7:
        date_delta = config.seven_days_ago
    elif days == 15:
        date_delta = config.fifteen_days_ago
    elif days == 30:
        date_delta = config.thirty_days_ago
    else:
        raise ValueError('Incorrect date range entered')

    new_cust_ws = openpyxl.Workbook()
    sheet = new_cust_ws.active
    rows = get_spreadsheet_data()
    values = get_cust_data_from_worksheet(rows, days, date_delta)
    # values is a list of cell values
    # each row looks like:
    # ([u'Michelle', u'Thomas', u'1-171-800-1954x38968'...,]

    # Write Headers to Spreadsheet
    sheet.cell(row=1, column=1).value = 'Name'
    sheet.cell(row=1, column=2).value = 'Last Name'
    sheet.cell(row=1, column=3).value = 'Phone Number'
    sheet.cell(row=1, column=4).value = 'Email Address'
    sheet.cell(row=1, column=5).value = 'Date'

    # TODO: Test that no records are left out.
    # TODO: Refactor the get_....data methods

    for row_num, row in enumerate(values, start=2):
        for col_num, cell in enumerate(row, start=1):
            sheet.cell(row=row_num, column=col_num).value =\
                values[row_num - 2][col_num - 1]

    new_cust_ws.save('cust-{0}-days.xlsx'.format(days))


if __name__ == '__main__':
    # Everything except the script name is passed to the main function.
    main(sys.argv[1:])
